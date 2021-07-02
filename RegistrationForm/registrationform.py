# Importing all the necessary libraries

import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# ----------------------------WINDOW---------------------------------------
# Defining window and setting its properties like title, geometry, background color, etc
window = tk.Tk()
window.title("Registration form")
window.geometry("450x360")
window.resizable(0, 0)
window.configure(bg="light green")
# --------------------------------------------------------------------------


# ------------------------------EXCEL----------------------------------------
# Setting up an excel workbook and defining its properties
wb = Workbook()
worksheet = wb.active
worksheet.title = "Entries"

# Excel worksheet heading and making it bold
details = ["Name", "Course", "Semester", "Roll Number", "Contact Number", "Email ID", "Address"]
worksheet.append(details)
for cols in range(1, len(details) + 1):
    worksheet[get_column_letter(cols) + '1'].font = Font(bold=True)

worksheet.column_dimensions['A'].width = 15
worksheet.column_dimensions['B'].width = 20
worksheet.column_dimensions['C'].width = 10
worksheet.column_dimensions['D'].width = 12
worksheet.column_dimensions['E'].width = 15
worksheet.column_dimensions['F'].width = 30
worksheet.column_dimensions['G'].width = 40
# ---------------------------------------------------------------------------


# ----------------------------------VALUES------------------------------------
# Type (string or int) of entries of the Entry widget and also setting its initial value
n_sem = tk.IntVar()
n_sem.set(0)
n_contact = tk.IntVar()
n_contact.set(0)

s_name = tk.StringVar()
s_name.set("")
s_course = tk.StringVar()
s_course.set("")
s_roll = tk.StringVar()
s_roll.set("")
s_mail = tk.StringVar()
s_mail.set("")
s_address = tk.StringVar()
s_address.set("")

# Function that sets all entry widget values to initial ""(for str receiving entry widget) or 0(for int receiving entry widgets) value
def set_to_initial():
    n_sem.set(0)
    n_contact.set(0)
    s_name.set("")
    s_course.set("")
    s_roll.set("")
    s_mail.set("")
    s_address.set("")
# -----------------------------------------------------------------------------


# ------------------------PROGRAM EXECUTION FUNCTION---------------------------
# Function to be executed on submitting the button.
def click_submit():
    # When all inputs are not entered
    if (s_name.get() == "" and s_course.get() == "" and n_sem.get() == "" and s_roll.get() == "" and n_contact.get() == "" and s_mail.get() == "" and s_address.get() == ""):
        messagebox.showerror("Empty Input", "Fields are empty and not filled properly")
    # When all entries are entered
    else:
        # Getting entries, appending to the worksheet, saving the workbook, resetting values to initial values
        entries = [s_name.get(), s_course.get(), n_sem.get(), s_roll.get(), n_contact.get(), s_mail.get(), s_address.get()]
        worksheet.append(entries)
        wb.save("Registration_Form_Entry.xlsx")
        set_to_initial()
# ---------------------------------------------------------------------


# ------------------------------WIDGETS---------------------------------
heading = tk.Label(master=window, text="Registration form", padx=3, pady=3, bg="light green", font=("Helvetica", 16, "bold"))
heading.grid(row=0, column=1, pady=5)

name = tk.Label(master=window, text="Name", padx=3, pady=3, bg="light green", font=("MS Sans Serif", 11, "bold"))
name.grid(row=1, column=0, pady=3, padx=3)
# Input receiving widget
nameinput = tk.Entry(master=window, width=45, textvariable=s_name)
nameinput.grid(row=1, column=1, pady=3, padx=5)

course = tk.Label(master=window, text="Course", padx=3, pady=3, bg="light green", font=("MS Sans Serif", 11, "bold"))
course.grid(row=2, column=0, pady=3, padx=3)
courseinput = tk.Entry(master=window, width=45, textvariable=s_course)
courseinput.grid(row=2, column=1, pady=3, padx=5)

sem = tk.Label(master=window, text="Sem", padx=3, pady=3, bg="light green", font=("MS Sans Serif", 11, "bold"))
sem.grid(row=3, column=0, pady=3, padx=3)
seminput = tk.Entry(master=window, width=45, textvariable=n_sem)
seminput.grid(row=3, column=1, pady=3, padx=5)

rollno = tk.Label(master=window, text="Roll No.", padx=3, pady=3, bg="light green", font=("MS Sans Serif", 11, "bold"))
rollno.grid(row=4, column=0, pady=3, padx=3)
rollnoinput = tk.Entry(master=window, width=45, textvariable=s_roll)
rollnoinput.grid(row=4, column=1, pady=3, padx=5)

contact = tk.Label(master=window, text="Contact No.", padx=3, pady=3, bg="light green", font=("MS Sans Serif", 11, "bold"))
contact.grid(row=5, column=0, pady=3, padx=3)
contactinput = tk.Entry(master=window, width=45, textvariable=n_contact)
contactinput.grid(row=5, column=1, pady=3, padx=5)

email = tk.Label(master=window, text="Email Id", padx=3, pady=3, bg="light green", font=("MS Sans Serif", 11, "bold"))
email.grid(row=6, column=0, pady=3, padx=3)
emailinput = tk.Entry(master=window, width=45, textvariable=s_mail)
emailinput.grid(row=6, column=1, pady=3, padx=5)

address = tk.Label(master=window, text="Address", padx=3, pady=3, bg="light green", font=("MS Sans Serif", 11, "bold"))
address.grid(row=7, column=0, pady=3, padx=3)
addressinput = tk.Entry(master=window, width=45, textvariable=s_address)
addressinput.grid(row=7, column=1, pady=3, padx=5)

# Button for submitting
submit = tk.Button(master=window, text="Submit", bg="green", fg="yellow", padx=5, pady=3, command=click_submit)
submit.grid(row=8, column=1, pady=5)

# Window quit button
end = tk.Button(master=window, text="END", bg="red", fg="black", padx=7, pady=1, command=window.quit)
end.grid(row=9, column=1, pady=10)
# -----------------------------------------------------------------------


# Executing the window
window.mainloop()